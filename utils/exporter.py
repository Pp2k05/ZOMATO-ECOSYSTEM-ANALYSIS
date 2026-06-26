# =============================================================================
#  utils/exporter.py  —  Export merged DataFrame to formatted Excel
# =============================================================================
import os
import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from config import COLUMNS


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
HEADER_BG   = "1F4E79"   # dark navy
HEADER_FG   = "FFFFFF"
ALT_ROW_BG  = "EBF3FB"   # light blue tint
ACCENT      = "2E75B6"

PLATFORM_COLORS = {
    "Reddit":          "FF5700",
    "Google Play":     "34A853",
    "Apple App Store": "555555",
    "Twitter":         "1DA1F2",
    "Quora":           "B92B27",
    "News":            "F4B400",
}


def _thin_border():
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def _style_header_row(ws, num_cols: int):
    fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    font = Font(name="Calibri", bold=True, color=HEADER_FG, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align


def _auto_col_widths(ws, col_max_widths: dict):
    """Set column widths, capped at per-column maximums."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        header_val = str(col_cells[0].value or "")
        max_len = max(
            len(header_val),
            max((len(str(c.value or "")) for c in col_cells[1:11]), default=0)  # sample 10 rows
        )
        cap = col_max_widths.get(header_val, 40)
        ws.column_dimensions[col_letter].width = min(max_len + 2, cap)


def export_to_excel(df: pd.DataFrame, output_path: str) -> None:
    """
    Write Sheet 1 (Raw Data) and Sheet 2 (Summary) to *output_path*.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    # Ensure every required column is present
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[COLUMNS].copy()

    # ---- Build summary -------------------------------------------------------
    summary = (
        df.groupby(["Platform", "Product_Tag"], sort=True)
          .size()
          .reset_index(name="Row_Count")
    )
    platform_totals = (
        df.groupby("Platform", sort=True)
          .size()
          .reset_index(name="Row_Count")
          .assign(Product_Tag="** TOTAL **")
    )
    summary = pd.concat([summary, platform_totals], ignore_index=True).sort_values(
        ["Platform", "Product_Tag"]
    )
    grand_total = pd.DataFrame([{
        "Platform": "** GRAND TOTAL **",
        "Product_Tag": "",
        "Row_Count": len(df),
    }])
    summary = pd.concat([summary, grand_total], ignore_index=True)

    # ---- Write to Excel -------------------------------------------------------
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Raw Data", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

        # ---- Style: Raw Data sheet -------------------------------------------
        ws_raw = writer.sheets["Raw Data"]
        ws_raw.freeze_panes = "A2"
        ws_raw.row_dimensions[1].height = 30

        _style_header_row(ws_raw, len(COLUMNS))

        col_caps = {
            "Post_ID": 20, "Platform": 18, "Source": 22, "Date": 14,
            "Username": 20, "Title": 45, "Text": 80, "Score": 12,
            "Product_Tag": 18, "URL": 55,
        }
        _auto_col_widths(ws_raw, col_caps)

        # Alternate row shading (cap at 10k rows to keep file size sane)
        alt_fill = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")
        wrap_cols = {"Title", "Text", "URL"}
        max_shade = min(ws_raw.max_row, 10001)
        for row_idx in range(2, max_shade + 1):
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                cell = ws_raw.cell(row=row_idx, column=col_idx)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                cell.alignment = Alignment(
                    wrap_text=(col_name in wrap_cols),
                    vertical="top"
                )

        # ---- Style: Summary sheet --------------------------------------------
        ws_sum = writer.sheets["Summary"]
        ws_sum.freeze_panes = "A2"
        ws_sum.row_dimensions[1].height = 28

        _style_header_row(ws_sum, 3)
        _auto_col_widths(ws_sum, {"Platform": 25, "Product_Tag": 22, "Row_Count": 14})

        # Colour-code Platform column
        for row_idx in range(2, ws_sum.max_row + 1):
            plat_cell = ws_sum.cell(row=row_idx, column=1)
            count_cell = ws_sum.cell(row=row_idx, column=3)
            plat_val = str(plat_cell.value or "")
            color = PLATFORM_COLORS.get(plat_val, ACCENT)
            plat_cell.font = Font(name="Calibri", bold="TOTAL" in plat_val or "GRAND" in plat_val, color=color)
            count_cell.alignment = Alignment(horizontal="center")
            if "GRAND" in plat_val:
                for c in (ws_sum.cell(row=row_idx, column=i) for i in range(1, 4)):
                    c.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                    c.font = Font(name="Calibri", bold=True, color="FFFFFF")

    print(f"\n✅  Exported {len(df):,} rows → {output_path}")
    print(f"    Sheet 1: Raw Data  |  Sheet 2: Summary")
